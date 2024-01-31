import csv
import pypdf

from openpyxl import load_workbook

from tests.conftest import PROJECT_RESOURCES_PATH


def test_file_xlsx_after_archivate():
    lw = load_workbook(PROJECT_RESOURCES_PATH + '/' + 'file_example_XLSX_50.xlsx')
    sheet = lw.active
    assert sheet.cell(1, 1).value == 0
    assert sheet.cell(1, 2).value == 'First Name'
    assert sheet.cell(1, 3).value == 'Last Name'


def test_file_csv_after_archivate():
    with open(PROJECT_RESOURCES_PATH + '/' + 'file-coderstool.csv', 'r') as f:
        reader = csv.reader(f)
        assert next(reader)[0] == '5GD1LBUQ15JTV7QN'


def test_file_pdf_after_archivate():
    reader = pypdf.PdfReader(PROJECT_RESOURCES_PATH + '/' + 'SQETestCaseSpecificationTemplate.pdf')
    assert reader.pages[0].extract_text().split('\n')[3].strip() == 'Test Case Specification Template'
